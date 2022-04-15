import openpyxl



def getRowCount(filename,sheetname):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)
def getColumnCount(filename,sheetname):
    workbook=openpyxl.load_workbook(filename)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)
def readData(filename,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value
def writeData(filename,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(filename)




